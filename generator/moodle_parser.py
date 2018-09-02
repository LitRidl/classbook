import html
import re
from bs4 import BeautifulSoup, Comment
import random
import json
from collections import OrderedDict, defaultdict
import sys
import os
from openpyxl import load_workbook
from os.path import join



def fmt(tag):
    if 'Инф: 4. Электр.таблицы/Информация вокруг нас (5-6)' in tag:
        return 'Обработка числовых данных в электр. таблицах'
    if 'Кл: ' in tag:
        tag += ' класс'
    if 'Ур: ' in tag:
        tag += ' уровень'
    for i in range(1, 8 + 1):
        tag = tag.replace('{}. '.format(i), '')
    return (tag
            .replace('Кл: ', '')
            .replace('Ур: ', '')
            .replace('Фин: ', '')
            .replace('Инф: ', ''))


def tags_classifier(question, fake_attributes=False):
    grades = [
        'Кл: 1. 5-6',
        'Кл: 2. 7-9',
        'Кл: 3. 10-11',
    ]
    difficulties = [
        'Ур: 1. Базовый',
        'Ур: 2. Повышенный',
        'Ур: 3. Высокий',
    ]
    topics_finances = [
        'Фин: 1. Расходы', 
        'Фин: 2. Доходы', 
        'Фин: 3. Семейный бюджет', 
        'Фин: 4. Сбережения и инвестиции', 
        'Фин: 5. Платежи и расчёты', 
        'Фин: 6. Кредиты и займы', 
        'Фин: 7. Страхование', 
        'Фин: 8. Риски и финансовая безопасность',
    ]
    topics_informatics = [
        'Инф: 1. Информация и информационные процессы', 
        'Инф: 2. Алгоритмизация и программирование', 
        'Инф: 3. Моделирование и формализация', 
        'Инф: 4. Электр.таблицы/Информация вокруг нас (5-6)', 
        'Инф: 6. Измерение количества информации', 
        'Инф: 7. Информационная безопасность', 
    ]
    task_types = [
        'Тип: 1. Автоматизированная', 
        'Тип: 2. Ручная', 
    ]
    tags = question['_tags']

    attributes = {
        'grade':             [fmt(random.choice(grades))] if fake_attributes else [],
        'difficulty':        fmt(random.choice(difficulties)) if fake_attributes else None,
        'topics_finances':    [fmt(random.choice(topics_finances))] if fake_attributes else [],
        'topics_informatics': [fmt(random.choice(topics_informatics))] if fake_attributes else [],
        'task_type':          [fmt(random.choice(task_types))] if fake_attributes else [],
        '_fake_attributes':  ['grade', 'difficulty', 'topics_finances', 'topics_informatics', 'task_type'],
    }

    for tag in tags:
        if tag in grades:
            attributes['grade'].append(fmt(tag))
            if 'grade' in attributes['_fake_attributes']:
                attributes['_fake_attributes'].remove('grade')
        elif tag in difficulties:
            attributes['difficulty'] = fmt(tag)
            if 'difficulty' in attributes['_fake_attributes']:
                attributes['_fake_attributes'].remove('difficulty')
            else:
                print('Question with multiple difficulties: question {}, {}, ({})'.format(question['code'], question['name'], [tag, attributes['difficulty']]))
        elif tag in topics_finances:
            attributes['topics_finances'].append(fmt(tag))
            if 'topics_finances' in attributes['_fake_attributes']:
                attributes['_fake_attributes'].remove('topics_finances')
        elif tag in topics_informatics:
            attributes['topics_informatics'].append(fmt(tag))
            if 'topics_informatics' in attributes['_fake_attributes']:
                attributes['_fake_attributes'].remove('topics_informatics')
        elif tag in task_types:
            attributes['task_type'] = fmt(tag)
            attributes['_fake_attributes'].remove('task_type')

    if not fake_attributes:
        if len(attributes['_fake_attributes']) > 0:
            print('Tag missing: question {}, {}, ({})'.format(question['code'], question['name'], attributes['_fake_attributes']))
            # raise Exception("Incomplete tags for attributes: {}".format(question))
        del attributes['_fake_attributes']

    if len(attributes['grade']) > 1:
        print('Question with multiple grades: question {}, {}, ({})'.format(question['code'], question['name'], attributes['grade']))

    return attributes


def parse_question(q, _id=None):
    name = q.find('name').text.strip()
    question = {
        'name': name.split(' ', 1)[1].strip(),
        'code': name.split(' ', 1)[0].strip(),
        'text': html.unescape(q.find('questiontext').find('text')
                              .prettify(formatter='xml').strip()
                              .lstrip('<text>').rstrip('</text>').strip()),
        'type': q.get('type').strip(),
        'answer': None,
        'answer_tolerance': None,
        '_tags': [t.find('text').text.strip() for t in q.find('tags').find_all('tag')]
    }
    if _id:
        question['_id'] = _id
    if question['type'] in ('shortanswer', 'numerical'):
        question['answer'] = q.find('answer').find('text').text.strip()
    if question['type'] == 'numerical':
        question['answer_tolerance'] = float(q.find('answer').find('tolerance').text.strip())
        question['answer'] = float(question['answer'])
    question.update(tags_classifier(question))

    return question


def parse_questions(xml):
    # xml = xml.replace('src="http://finformatika.ru/images/', 'src="/images/')
    # xml = xml.replace('src="/images/', 'src="assets/finformatika.ru/images/')
    # xml.replace('href="http://finformatika.ru/files', 'href="/attachments')
    xml = xml.replace('href="http://finformatika.ru/', 'download href="assets/finformatika.ru/')
    xml = xml.replace('src="http://finformatika.ru/', 'src="assets/finformatika.ru/')
    soup = BeautifulSoup(xml, 'xml')

    soup_comments = [c.strip()for c in soup.findAll(text=lambda text: isinstance(text, Comment))
                     if 'question' in c]
    question_ids = [int(c.replace('question: ', ''))
                    for c in soup_comments if 'question: 0' not in c]

    # essay, numerical, shortanswer
    soup_questions = soup.find_all('question', type=re.compile('essay|numerical|shortanswer'))
    questions = [parse_question(q, _id=i)
                 for i, q in enumerate(soup_questions)]
    questions = [dict(q, question_id=q_id)
                 for q_id, q in zip(question_ids, questions)]

    # TREE LOGICS
    # codes = defaultdict(list)
    # for q in questions:
    #     codes[q['code']].append(q)
    # for k, v in codes.items():
    #     pid = [q['question_id'] for q in v if q['type'] == 'essay']
    #     for q in v:
    #         if q['type'] != 'essay':
    #             q['parent_id'] = pid

    return questions


def order_dict(d, order):
    return [OrderedDict(sorted(el.items(), key=lambda el: order.index(el[0]))) for el in d]



def list_excel_files(only_names=False):
    filenames = []
    for root, _, files in os.walk('../assets/files/excel/'):
        filenames += [join(root, name) for name in files if name.endswith('.xlsx')]
    return [os.path.basename(fn) for fn in filenames] if only_names else filenames


def filename_to_code(fname):
    fname = os.path.basename(fname).rstrip('.xlsx')
    return '{}.{}.{}.{}'.format(fname[0], fname[1], fname[2], fname[3:])


def is_excel_interactive(fname):
    wb = load_workbook(fname)
    ws_name = None
    for ws in wb.sheetnames:
        if '(решение)' in ws:
            ws_name = ws
    return ws_name



def questions_from_file(f):
    questions = parse_questions(open(f).read())
    order = ['_id', 'question_id', 'grade', 'difficulty', 'topics_finances',
             'topics_informatics', 'name', 'text', 'type', 'answer', 'answer_tolerance', 'code', 'task_type', 'excel_table_name', '_tags', '_fake_attributes']

    fnames = list_excel_files()
    fnames_used = []
    for q in questions:
        fname = None
        for f in fnames:
            if filename_to_code(f) == q['code']:
                fname = f
                break
        if fname and is_excel_interactive(fname) and q['type'] == 'essay':
            fnames_used.append(fname)
            q['excel_table_name'] = is_excel_interactive(fname)
            q['task_type'] = 'Тип: Excel-задача'
            q['type'] = 'excel'
        else:
            q['excel_table_name'] = None

    qs = order_dict(questions, order)

    # fnames.extend(fnames_used)
    # fnames.sort()
    # print(fnames)
    # print('Fnames: {}'.format(len(fnames)))
    # print('Qs: {}'.format(len([q for q in qs if q['excel_table_name']])))
    # print('Tts: {}'.format(len([q for q in qs if q['task_type'] == 'Тип: Excel-задача'])))
    # print('Ts: {}'.format(len([q for q in qs if q['type'] == 'excel'])))

    idx = 1
    for q in qs:
        res = q['text']
        if 'ftn1' in q['text']:
            res = res.replace('ftn1', 'ftn{}'.format(idx))
            res = res.replace('ftnref1', 'ftnref{}'.format(idx))
            idx += 1
            if 'ftn2' in q['text']:
                res = res.replace('ftn2', 'ftn{}'.format(idx))
                res = res.replace('ftnref2', 'ftnref{}'.format(idx))
                idx += 1
            q['text'] = res

    return qs


def parse_glossary_entry(e):
    return {
        'concept': e.find('CONCEPT').text.strip(),
        'definition': html.unescape(e.find('DEFINITION').prettify(formatter='xml').strip()
                                     .lstrip('<DEFINITION>').rstrip('</DEFINITION>').strip()),
    }


def glossary_from_file(xml_file):
    soup = BeautifulSoup(open(xml_file).read(), 'xml')
    return [parse_glossary_entry(e) for e in soup.find_all('ENTRY')]


def json_to_file(data, fname, indent=4):
    with open(fname, 'w') as f:
        json.dump(data, f, indent=indent, ensure_ascii=False)


def gen_index(qs):
    import bisect
    idx = defaultdict(lambda: defaultdict(list))
    fields = ['grade', 'type', 'difficulty', 'topics_finances', 'topics_informatics', 'task_type', 'excel_table_name']
    for q in qs:
        for field in fields:
            if isinstance(q[field], list):
                for v in q[field]:
                    bisect.insort(idx[field][v], q['question_id'])
            else:
                v = q[field]
                if field == 'type':
                    if v == 'excel':
                        v = 'Excel'
                    else:
                        v = 'Текст' if v == 'essay' else 'Интерактив'
                bisect.insort(idx[field][v], q['question_id'])
    return idx


if __name__ == '__main__':
    print('Outputting questions to {}'.format('moodle_data.json'))
    qs = questions_from_file('moodle_data/moodle_export_1.xml')

    # tags = sum((q['_tags'] for q in qs), [])
    # tags = list(set(tags))
    # tags.sort()
    # print(tags)

    json_to_file(qs, 'moodle_data.json')

    #print('Outputting index file to {}'.format('questions_index.json'))
    #json_to_file(gen_index(qs), 'questions_index.json', indent=None)

    # for q in qs:
    # # Block for visual error searching
    #     imgs = bs.find_all('a')
    #     for img in imgs:
    #         print(img)
    #     links = bs.find_all(href=True)
    #     for link in links:
    #         if 'assets' in link['href']:
    #             print(link)
    #     for link in links:
    #         href = link['href']
    #         if 'assets' in link['href']:
    #             try:
    #                 open('../' + href).close()
    #             except Exception as e:
    #                 print(e)
    # # Auto-search
    #     if q['code'].count('.') >= 4 and q['code'][-2] != '0':
    #         print('No leading zero + 4 x .:', q['code'])
    #     if q['code'].count('.') >= 4 and q['code'][-2] == '0' and (q['type'] == 'essay' or q['task_type'] == 'Тип: Ручная'):
    #         print('Non-auto and 4 x .:', q['code'], q['type'], q['task_type'])
    #     if q['code'].count('.') >= 4:
    #         pre = '.'.join(q['code'].split('.')[:4])
    #         if all(not pre in qq['code'] for qq in qs):
    #             print('No parent task:', q['code'])
    #     if len(q['grade']) != 1:
    #         print('Wrong grades qty: ', q['code'], q['grade'])
    #     if len(q['topics_finances']) != 1:
    #         print('Wrong topics_finances qty: ', q['code'], q['topics_finances'])
    #     if len(q['topics_informatics']) != 1:
    #         print('Wrong topics_informatics qty: ', q['code'], q['topics_informatics'])
    #     if '1drv' in q['text']:
    #         print('1drv detected: {code} {name}'.format(**q))
    #     bs = BeautifulSoup(q['text'], 'lxml')


    # for q in qs:
    #     if not q['code'][0].isdigit() or not q['code'][len(q['code'])-1].isdigit():
    #         print(q['code'], q['name'])
    #     if not q['name'][0].isalpha():
    #         print(q['code'], q['name'])
    #     if 'задача' in q['name'].lower():
    #         print(q['code'], q['name'])
    #     if not all(s.isdigit() or s == '.' for s in q['code']):
    #          print(q['code'], q['name'])

    # print('Outputting glossary to {}'.format('glossary.json'))
    # glossary_file = 'moodle_data/glossary_data_1.xml'
    # glossary = glossary_from_file(glossary_file)
    # json_to_file(glossary, 'glossary.json')
